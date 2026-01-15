# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# 스타일 정의
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

category_font = Font(bold=True, size=12)
category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 테이블 데이터 정의
tables = {
    "사용자": [
        {
            "name": "user_roles",
            "desc": "유저 등급 테이블",
            "columns": [
                ("code", "VARCHAR(10)", "NO", "-", "PK", "역할 코드"),
                ("role", "VARCHAR(50)", "NO", "-", "-", "역할명"),
            ]
        },
        {
            "name": "countries",
            "desc": "국가 테이블",
            "columns": [
                ("code", "VARCHAR(10)", "NO", "-", "PK", "국가 코드"),
                ("name", "VARCHAR(50)", "NO", "-", "-", "국가명"),
                ("currency", "VARCHAR(10)", "NO", "USD", "-", "결제 통화"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "users",
            "desc": "유저 테이블",
            "columns": [
                ("email", "VARCHAR(255)", "NO", "-", "PK", "이메일 (로그인 ID)"),
                ("password_hash", "VARCHAR(255)", "YES", "-", "-", "비밀번호 해시"),
                ("roles_code", "VARCHAR(10)", "NO", "002", "FK", "역할 코드"),
                ("name", "VARCHAR(100)", "YES", "-", "-", "이름"),
                ("phone", "VARCHAR(20)", "YES", "-", "-", "전화번호"),
                ("country_code", "VARCHAR(10)", "YES", "KR", "FK", "국가 코드"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "계정 활성화 상태"),
                ("deactivated_at", "TIMESTAMP", "YES", "-", "-", "계정 비활성화 시점"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "user_social_accounts",
            "desc": "소셜 계정 연동 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email", "VARCHAR(255)", "NO", "-", "FK", "사용자 이메일"),
                ("provider", "VARCHAR(20)", "NO", "-", "-", "소셜 제공자"),
                ("provider_id", "VARCHAR(255)", "NO", "-", "-", "소셜 플랫폼 사용자 ID"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "email_verifications",
            "desc": "이메일 인증 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("email", "VARCHAR(255)", "NO", "-", "UNIQUE", "인증할 이메일"),
                ("verification_code", "VARCHAR(6)", "NO", "-", "-", "6자리 인증 코드"),
                ("expires_at", "TIMESTAMP", "NO", "-", "-", "코드 만료 시간"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "password_reset_tokens",
            "desc": "비밀번호 재설정 토큰 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("email", "VARCHAR(255)", "NO", "-", "UNIQUE", "재설정할 계정 이메일"),
                ("reset_code", "VARCHAR(6)", "NO", "-", "-", "6자리 인증 코드"),
                ("expires_at", "TIMESTAMP", "NO", "-", "-", "코드 만료 시간"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "token_blacklist",
            "desc": "토큰 블랙리스트 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("token", "VARCHAR(500)", "NO", "-", "-", "블랙리스트 JWT 토큰"),
                ("user_email", "VARCHAR(255)", "NO", "-", "-", "토큰 소유자 이메일"),
                ("expires_at", "TIMESTAMP", "NO", "-", "-", "토큰 만료 시간"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
    ],
    "결제": [
        {
            "name": "products",
            "desc": "상품 종류 테이블",
            "columns": [
                ("code", "VARCHAR(3)", "NO", "-", "PK", "상품 코드 (000~999)"),
                ("name", "VARCHAR(255)", "NO", "-", "-", "상품명"),
                ("description", "TEXT", "YES", "-", "-", "상품 설명"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "활성화 여부"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "price_plans",
            "desc": "상품 가격 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("product_code", "VARCHAR(3)", "NO", "-", "FK", "상품 코드"),
                ("name", "VARCHAR(100)", "NO", "-", "-", "요금제명"),
                ("description", "VARCHAR(100)", "YES", "-", "-", "요금제 설명"),
                ("price", "DECIMAL(18,2)", "NO", "-", "-", "가격"),
                ("currency", "VARCHAR(10)", "NO", "KRW", "-", "통화"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "활성화 여부"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "promotions",
            "desc": "프로모션/쿠폰 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("code", "VARCHAR(50)", "NO", "-", "UNIQUE", "프로모션 코드"),
                ("name", "VARCHAR(100)", "NO", "-", "-", "프로모션명"),
                ("discount_type", "INTEGER", "NO", "-", "-", "할인 유형 (퍼센트)"),
                ("discount_value", "DECIMAL(18,2)", "NO", "-", "-", "할인 값"),
                ("product_code", "VARCHAR(3)", "YES", "-", "FK", "특정 상품에만 적용"),
                ("usage_limit", "INTEGER", "YES", "-", "-", "사용 횟수 제한"),
                ("usage_count", "INTEGER", "NO", "0", "-", "현재 사용 횟수"),
                ("valid_from", "TIMESTAMP", "NO", "NOW", "-", "유효 시작일"),
                ("valid_until", "TIMESTAMP", "YES", "-", "-", "유효 종료일"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "활성화 여부"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "subscriptions",
            "desc": "유저 구독 관리 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email", "VARCHAR(255)", "YES", "-", "FK", "사용자 이메일"),
                ("product_code", "VARCHAR(3)", "NO", "-", "FK", "상품 코드"),
                ("price_plan_id", "BIGINT", "NO", "-", "FK", "요금제 ID"),
                ("status", "VARCHAR(1)", "NO", "A", "-", "상태 (A/E/C)"),
                ("start_date", "TIMESTAMP", "NO", "-", "-", "구독 시작일"),
                ("end_date", "TIMESTAMP", "NO", "-", "-", "구독 종료일"),
                ("auto_renew", "BOOLEAN", "NO", "FALSE", "-", "자동 갱신 여부"),
                ("billing_key_id", "BIGINT", "YES", "-", "FK", "빌링키 ID"),
                ("next_billing_date", "TIMESTAMP", "YES", "-", "-", "다음 결제 예정일"),
                ("billing_cycle", "VARCHAR(20)", "YES", "YEARLY", "-", "결제 주기"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "billing_keys",
            "desc": "빌링키 테이블 (자동결제용)",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email", "VARCHAR(255)", "NO", "-", "FK", "사용자 이메일"),
                ("billing_key", "VARCHAR(255)", "NO", "-", "-", "토스페이먼츠 빌링키"),
                ("customer_key", "VARCHAR(255)", "NO", "-", "-", "고객 식별키"),
                ("card_company", "VARCHAR(50)", "YES", "-", "-", "카드사명"),
                ("card_number", "VARCHAR(20)", "YES", "-", "-", "마스킹된 카드번호"),
                ("card_type", "VARCHAR(20)", "YES", "-", "-", "카드 유형"),
                ("owner_type", "VARCHAR(20)", "YES", "-", "-", "소유자 유형"),
                ("is_default", "BOOLEAN", "NO", "FALSE", "-", "기본 결제 수단"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "활성화 여부"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "subscription_payments",
            "desc": "구독 결제 이력 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("subscription_id", "BIGINT", "NO", "-", "FK", "구독 ID"),
                ("billing_key_id", "BIGINT", "YES", "-", "FK", "빌링키 ID"),
                ("payment_key", "VARCHAR(255)", "YES", "-", "-", "결제 키"),
                ("order_id", "VARCHAR(255)", "NO", "-", "-", "주문 ID"),
                ("amount", "DECIMAL(15,2)", "NO", "-", "-", "결제 금액"),
                ("status", "VARCHAR(20)", "NO", "PENDING", "-", "결제 상태"),
                ("billing_date", "DATE", "NO", "-", "-", "결제일"),
                ("paid_at", "TIMESTAMP", "YES", "-", "-", "결제 완료 시간"),
                ("failure_reason", "TEXT", "YES", "-", "-", "실패 사유"),
                ("retry_count", "INT", "NO", "0", "-", "재시도 횟수"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "payments",
            "desc": "결제 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email_fk", "VARCHAR(255)", "YES", "-", "FK", "사용자 이메일 (FK)"),
                ("user_email", "VARCHAR(255)", "NO", "-", "-", "사용자 이메일"),
                ("user_name", "VARCHAR(100)", "YES", "-", "-", "사용자 이름"),
                ("subscription_id", "BIGINT", "YES", "-", "FK", "구독 ID"),
                ("price_plan_id", "BIGINT", "YES", "-", "FK", "요금제 ID"),
                ("amount", "DECIMAL(18,2)", "NO", "-", "-", "결제 금액"),
                ("currency", "VARCHAR(10)", "NO", "KRW", "-", "통화"),
                ("status", "VARCHAR(1)", "NO", "P", "-", "상태 (P/C/F/R)"),
                ("order_name", "VARCHAR(255)", "YES", "-", "-", "주문명"),
                ("paid_at", "TIMESTAMP", "YES", "-", "-", "결제 완료 시간"),
                ("refunded_at", "TIMESTAMP", "YES", "-", "-", "환불 시간"),
                ("refund_amount", "DECIMAL(18,2)", "YES", "-", "-", "환불 금액"),
                ("refund_reason", "TEXT", "YES", "-", "-", "환불 사유"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "payment_details",
            "desc": "결제 상세 테이블",
            "columns": [
                ("payment_id", "BIGINT", "NO", "-", "PK,FK", "결제 ID"),
                ("payment_method", "VARCHAR(50)", "YES", "-", "-", "결제 수단"),
                ("payment_provider", "VARCHAR(50)", "YES", "-", "-", "PG사"),
                ("order_id", "VARCHAR(100)", "YES", "-", "-", "주문 ID"),
                ("payment_key", "VARCHAR(255)", "YES", "-", "-", "결제 키"),
                ("card_company", "VARCHAR(50)", "YES", "-", "-", "카드사명"),
                ("card_number", "VARCHAR(50)", "YES", "-", "-", "마스킹된 카드번호"),
                ("installment_months", "INT", "YES", "-", "-", "할부 개월수"),
                ("approve_no", "VARCHAR(50)", "YES", "-", "-", "카드 승인번호"),
                ("easy_pay_provider", "VARCHAR(50)", "YES", "-", "-", "간편결제 제공자"),
                ("bank_code", "VARCHAR(20)", "YES", "-", "-", "은행 코드"),
                ("bank_name", "VARCHAR(50)", "YES", "-", "-", "은행명"),
                ("account_number", "VARCHAR(50)", "YES", "-", "-", "가상계좌 번호"),
                ("due_date", "TIMESTAMP", "YES", "-", "-", "입금 기한"),
                ("depositor_name", "VARCHAR(100)", "YES", "-", "-", "입금자명"),
                ("settlement_status", "VARCHAR(20)", "YES", "-", "-", "정산 상태"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
    ],
    "로그": [
        {
            "name": "activity_logs",
            "desc": "활동 로그 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email", "VARCHAR(255)", "YES", "-", "FK", "사용자 이메일"),
                ("action", "VARCHAR(50)", "NO", "-", "-", "활동 유형"),
                ("target_type", "VARCHAR(50)", "YES", "-", "-", "대상 타입"),
                ("target_id", "BIGINT", "YES", "-", "-", "대상 ID"),
                ("description", "TEXT", "YES", "-", "-", "상세 설명"),
                ("ip_address", "VARCHAR(50)", "YES", "-", "-", "IP 주소"),
                ("user_agent", "TEXT", "YES", "-", "-", "브라우저 정보"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "user_change_logs",
            "desc": "유저 정보 변경 로그 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("user_email", "VARCHAR(255)", "YES", "-", "-", "변경된 사용자 이메일"),
                ("changed_field", "VARCHAR(50)", "NO", "-", "-", "변경된 필드명"),
                ("old_value", "TEXT", "YES", "-", "-", "이전 값"),
                ("new_value", "TEXT", "YES", "-", "-", "새 값"),
                ("changed_by_email", "VARCHAR(255)", "YES", "-", "-", "변경한 사람"),
                ("change_reason", "TEXT", "YES", "-", "-", "변경 사유"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
        {
            "name": "admin_logs",
            "desc": "관리자 로그 테이블",
            "columns": [
                ("id", "BIGINT", "NO", "AUTO", "PK", "기본키"),
                ("admin_email", "VARCHAR(255)", "NO", "-", "FK", "관리자 이메일"),
                ("action", "VARCHAR(100)", "NO", "-", "-", "작업 유형"),
                ("target_type", "VARCHAR(50)", "NO", "-", "-", "대상 타입"),
                ("target_id", "BIGINT", "YES", "-", "-", "대상 ID"),
                ("description", "TEXT", "YES", "-", "-", "작업 설명"),
                ("ip_address", "VARCHAR(50)", "YES", "-", "-", "IP 주소"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
            ]
        },
    ],
    "라이선스": [
        {
            "name": "license_plans",
            "desc": "라이선스 플랜 테이블",
            "columns": [
                ("id", "UUID", "NO", "uuid_v4()", "PK", "기본키"),
                ("product_id", "UUID", "NO", "-", "-", "상품 ID"),
                ("code", "VARCHAR(64)", "NO", "-", "-", "플랜 코드"),
                ("name", "VARCHAR(255)", "NO", "-", "-", "플랜명"),
                ("description", "TEXT", "YES", "-", "-", "플랜 설명"),
                ("license_type", "VARCHAR(32)", "NO", "-", "-", "라이선스 유형"),
                ("duration_days", "INT", "NO", "-", "-", "유효기간 (일)"),
                ("grace_days", "INT", "NO", "0", "-", "유예기간 (일)"),
                ("max_activations", "INT", "NO", "1", "-", "최대 기기 수"),
                ("max_concurrent_sessions", "INT", "NO", "1", "-", "최대 동시 세션"),
                ("allow_offline_days", "INT", "NO", "0", "-", "오프라인 허용 일수"),
                ("is_active", "BOOLEAN", "NO", "TRUE", "-", "활성화 여부"),
                ("is_deleted", "BOOLEAN", "NO", "FALSE", "-", "삭제 여부"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "license_plan_entitlements",
            "desc": "플랜 기능 권한 테이블",
            "columns": [
                ("id", "UUID", "NO", "uuid_v4()", "PK", "기본키"),
                ("plan_id", "UUID", "NO", "-", "FK", "플랜 ID"),
                ("entitlement_key", "VARCHAR(100)", "NO", "-", "-", "기능 식별자"),
            ]
        },
        {
            "name": "licenses",
            "desc": "라이선스 테이블",
            "columns": [
                ("id", "UUID", "NO", "uuid_v4()", "PK", "기본키"),
                ("owner_type", "VARCHAR(20)", "NO", "-", "-", "소유자 유형"),
                ("owner_id", "UUID", "NO", "-", "-", "소유자 ID"),
                ("product_id", "UUID", "NO", "-", "-", "상품 ID"),
                ("plan_id", "UUID", "YES", "-", "FK", "플랜 ID"),
                ("license_type", "VARCHAR(20)", "NO", "-", "-", "라이선스 유형"),
                ("usage_category", "VARCHAR(30)", "NO", "-", "-", "사용 용도"),
                ("status", "VARCHAR(20)", "NO", "-", "-", "라이선스 상태"),
                ("issued_at", "TIMESTAMP", "NO", "NOW", "-", "발급일시"),
                ("valid_from", "TIMESTAMP", "NO", "NOW", "-", "유효 시작일"),
                ("valid_until", "TIMESTAMP", "YES", "-", "-", "유효 종료일"),
                ("policy_snapshot", "JSONB", "YES", "-", "-", "정책 스냅샷"),
                ("license_key", "VARCHAR(50)", "YES", "-", "UNIQUE", "라이선스 키"),
                ("source_order_id", "UUID", "YES", "-", "-", "결제 주문 ID"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "license_activations",
            "desc": "라이선스 활성화 테이블",
            "columns": [
                ("id", "UUID", "NO", "uuid_v4()", "PK", "기본키"),
                ("license_id", "UUID", "NO", "-", "FK", "라이선스 ID"),
                ("device_fingerprint", "VARCHAR(255)", "NO", "-", "-", "기기 식별 해시"),
                ("status", "VARCHAR(20)", "NO", "-", "-", "활성화 상태"),
                ("activated_at", "TIMESTAMP", "NO", "NOW", "-", "활성화 일시"),
                ("last_seen_at", "TIMESTAMP", "NO", "NOW", "-", "마지막 확인 일시"),
                ("client_version", "VARCHAR(50)", "YES", "-", "-", "클라이언트 버전"),
                ("client_os", "VARCHAR(100)", "YES", "-", "-", "클라이언트 OS"),
                ("last_ip", "VARCHAR(45)", "YES", "-", "-", "마지막 IP"),
                ("device_display_name", "VARCHAR(100)", "YES", "-", "-", "기기 표시명"),
                ("deactivated_at", "TIMESTAMP", "YES", "-", "-", "비활성화 일시"),
                ("deactivated_reason", "VARCHAR(50)", "YES", "-", "-", "비활성화 사유"),
                ("offline_token", "VARCHAR(2000)", "YES", "-", "-", "오프라인 토큰"),
                ("offline_token_expires_at", "TIMESTAMP", "YES", "-", "-", "오프라인 토큰 만료"),
                ("created_at", "TIMESTAMP", "NO", "NOW", "-", "생성일시"),
                ("updated_at", "TIMESTAMP", "NO", "NOW", "-", "수정일시"),
            ]
        },
        {
            "name": "revoked_offline_tokens",
            "desc": "무효화된 오프라인 토큰 테이블",
            "columns": [
                ("id", "UUID", "NO", "uuid_v4()", "PK", "기본키"),
                ("license_id", "UUID", "NO", "-", "FK", "라이선스 ID"),
                ("activation_id", "UUID", "YES", "-", "-", "활성화 ID"),
                ("device_fingerprint", "VARCHAR(255)", "YES", "-", "-", "기기 식별 해시"),
                ("token_hash", "VARCHAR(255)", "NO", "-", "-", "토큰 해시"),
                ("revoked_at", "TIMESTAMP", "NO", "NOW", "-", "무효화 일시"),
                ("reason", "VARCHAR(255)", "YES", "-", "-", "무효화 사유"),
            ]
        },
    ],
}

# 시트 1: 테이블 목록
ws1 = wb.active
ws1.title = "테이블 목록"

# 헤더
headers = ["No", "분류", "테이블명", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 데이터
row_num = 2
table_num = 1
for category, table_list in tables.items():
    for table in table_list:
        ws1.cell(row=row_num, column=1, value=table_num).border = thin_border
        ws1.cell(row=row_num, column=1).alignment = center_alignment
        ws1.cell(row=row_num, column=2, value=category).border = thin_border
        ws1.cell(row=row_num, column=3, value=table["name"]).border = thin_border
        ws1.cell(row=row_num, column=4, value=table["desc"]).border = thin_border
        row_num += 1
        table_num += 1

# 열 너비 조정
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 35

# 시트 2: 테이블 상세 정의
ws2 = wb.create_sheet("테이블 상세 정의")

row_num = 1
for category, table_list in tables.items():
    # 카테고리 헤더
    cell = ws2.cell(row=row_num, column=1, value=f"[{category}]")
    cell.font = category_font
    cell.fill = category_fill
    ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    row_num += 1

    for table in table_list:
        # 테이블명
        cell = ws2.cell(row=row_num, column=1, value=f"{table['name']} - {table['desc']}")
        cell.font = Font(bold=True, size=11)
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        row_num += 1

        # 컬럼 헤더
        col_headers = ["No", "컬럼명", "데이터 타입", "NULL", "기본값", "PK/FK", "설명"]
        for col, header in enumerate(col_headers, 1):
            cell = ws2.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        row_num += 1

        # 컬럼 데이터
        for idx, col_data in enumerate(table["columns"], 1):
            ws2.cell(row=row_num, column=1, value=idx).border = thin_border
            ws2.cell(row=row_num, column=1).alignment = center_alignment
            for col, value in enumerate(col_data, 2):
                cell = ws2.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col <= 6:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
            row_num += 1

        row_num += 1  # 테이블 사이 빈 줄

# 열 너비 조정
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 35

# 저장
wb.save("테이블정의서.xlsx")
print("엑셀 파일이 생성되었습니다: 테이블정의서.xlsx")
