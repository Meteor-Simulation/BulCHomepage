#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
테이블정의서.md를 엑셀 파일로 변환하는 스크립트
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def parse_markdown_table(lines):
    """마크다운 테이블을 파싱하여 리스트로 반환"""
    rows = []
    for line in lines:
        line = line.strip()
        if line.startswith('|') and not line.startswith('|--') and not line.startswith('|-'):
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if cells and not all(c.replace('-', '') == '' for c in cells):
                rows.append(cells)
    return rows

def create_excel_from_md(md_path, output_path):
    """마크다운 파일을 읽어 엑셀로 변환"""

    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 목록 시트 생성
    ws_index = wb.active
    ws_index.title = '테이블 목록'

    # 테이블 목록 파싱
    index_match = re.search(r'## 테이블 목록\s*\n([\s\S]*?)(?=\n---|\n## \d)', content)
    if index_match:
        table_lines = index_match.group(1).strip().split('\n')
        table_data = parse_markdown_table(table_lines)

        for row_idx, row in enumerate(table_data, 1):
            for col_idx, cell in enumerate(row, 1):
                c = ws_index.cell(row=row_idx, column=col_idx, value=cell)
                c.border = thin_border
                if row_idx == 1:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = header_alignment
                else:
                    c.font = cell_font
                    c.alignment = center_alignment if col_idx == 1 else cell_alignment

        # 열 너비 조정
        ws_index.column_dimensions['A'].width = 6
        ws_index.column_dimensions['B'].width = 30
        ws_index.column_dimensions['C'].width = 35
        ws_index.column_dimensions['D'].width = 12

    # 각 테이블 정의 파싱
    table_pattern = r'## (\d+)\. (\w+) \(([^)]+)\)\s*\n\n([^\n]+)\s*\n\n\| No \|([\s\S]*?)(?=\n---|\n## \d|\n\*\*[가-힣]|\Z)'

    for match in re.finditer(table_pattern, content):
        table_num = match.group(1)
        table_name = match.group(2)
        table_desc_kr = match.group(3)
        table_description = match.group(4).strip()
        table_content = '| No |' + match.group(5)

        # 시트 이름 (31자 제한)
        sheet_name = f"{table_num}.{table_name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 테이블 설명
        ws.cell(row=1, column=1, value=f"{table_name} ({table_desc_kr})")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws.cell(row=2, column=1, value=table_description)
        ws.cell(row=2, column=1).font = Font(size=10, color='666666')
        ws.merge_cells('A2:F2')

        # 테이블 데이터
        table_lines = table_content.strip().split('\n')
        table_data = parse_markdown_table(table_lines)

        start_row = 4
        for row_idx, row in enumerate(table_data):
            for col_idx, cell in enumerate(row, 1):
                c = ws.cell(row=start_row + row_idx, column=col_idx, value=cell)
                c.border = thin_border
                if row_idx == 0:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = header_alignment
                else:
                    c.font = cell_font
                    if col_idx == 1:  # No 컬럼
                        c.alignment = center_alignment
                    elif col_idx in [4, 5]:  # NULL, 기본값
                        c.alignment = center_alignment
                    else:
                        c.alignment = cell_alignment

        # 열 너비 조정
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 45

    # 저장
    wb.save(output_path)
    print(f"엑셀 파일 생성 완료: {output_path}")

if __name__ == '__main__':
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    md_path = os.path.join(script_dir, '테이블정의서.md')
    output_path = os.path.join(script_dir, '테이블정의서.xlsx')

    create_excel_from_md(md_path, output_path)
